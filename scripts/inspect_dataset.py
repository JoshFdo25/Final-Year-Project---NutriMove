"""Inspect all Sri Lanka diet reference XLSX files to understand structure."""
import openpyxl
import os

BASE = r"C:\Users\joshw\OneDrive\Desktop\FYP_Work\datasets\osfstorage-archive\Diet Reference Tables for Sri Lanka"
OUT = r"C:\Users\joshw\OneDrive\Desktop\FYP_Work\inspect_output.txt"

files = [
    "Sri Lanka Food Composition Table_20240514.xlsx",
    "Sri Lanka Food Group Table_20240506.xlsx",
    "Sri Lanka Portion Conversion Estimation Table_20240418.xlsx",
    "Sri Lanka Standard Recipes Database Ingredients FCT_20240419.xlsx",
    "Sri Lanka Standard Recipes Database_20240514.xlsx",
]

with open(OUT, "w", encoding="utf-8") as f:
    for fname in files:
        path = os.path.join(BASE, fname)
        f.write(f"\n{'='*70}\n")
        f.write(f"FILE: {fname}\n")
        f.write(f"{'='*70}\n")
        
        wb = openpyxl.load_workbook(path, data_only=True)
        f.write(f"Sheets: {wb.sheetnames}\n")
        
        for sname in wb.sheetnames:
            ws = wb[sname]
            f.write(f"\n  --- Sheet: '{sname}' ({ws.max_row} rows x {ws.max_column} cols) ---\n")
            
            # Print first 5 rows
            for r in range(1, min(6, ws.max_row + 1)):
                vals = []
                for c in range(1, min(ws.max_column + 1, 40)):
                    v = ws.cell(r, c).value
                    if v is not None:
                        vals.append(f"C{c}={v}")
                f.write(f"    Row {r}: {', '.join(vals)}\n")
        
        wb.close()

    f.write("\n\nDONE - All files inspected.\n")

print("Output written to inspect_output.txt")
