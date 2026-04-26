"""
Phase 3 — Food Database Processor (v2 — Fixed)
Fixes: 1) Better food group classification using name-based approach
       2) Precise alcohol filtering (no false positives for ginger, drumstick, etc.)
"""

import openpyxl
import json
import os
import re

BASE = r"C:\Users\joshw\OneDrive\Desktop\FYP_Work\datasets\osfstorage-archive\Diet Reference Tables for Sri Lanka"
OUT_DIR = r"C:\Users\joshw\OneDrive\Desktop\FYP_Work\diet_planner_app\assets\data"
os.makedirs(OUT_DIR, exist_ok=True)

def safe_float(val):
    if val is None: return 0.0
    try: return float(val)
    except: return 0.0

# ─────────────────────────────────────────────────────────
# 1. PARSE FOOD COMPOSITION TABLE (FCT)
# ─────────────────────────────────────────────────────────
print("=" * 60)
print("Step 1: Parsing Food Composition Table...")
fct_path = os.path.join(BASE, "Sri Lanka Food Composition Table_20240514.xlsx")
wb = openpyxl.load_workbook(fct_path, data_only=True)
ws = wb["FCT"]

foods = {}
for row in range(2, ws.max_row + 1):
    source    = ws.cell(row, 1).value or ""
    food_code = ws.cell(row, 2).value
    food_name = ws.cell(row, 3).value
    food_type = ws.cell(row, 4).value

    if not food_code or not food_name:
        continue

    food_code = str(food_code).strip()
    food_name = str(food_name).strip()

    calories = safe_float(ws.cell(row, 5).value)
    protein  = safe_float(ws.cell(row, 6).value)
    fat      = safe_float(ws.cell(row, 7).value)
    carbs    = safe_float(ws.cell(row, 8).value)
    fiber    = safe_float(ws.cell(row, 9).value)

    if calories <= 0:
        continue

    foods[food_code] = {
        "id": food_code,
        "name": food_name,
        "source": str(source).strip(),
        "region": "sri_lanka",
        "food_type": food_type if food_type else "S",
        "food_group": "",
        "meal_types": [],
        "dietary_tags": [],
        "recommendable": True,
        "per_100g": {
            "calories": round(calories, 1),
            "protein": round(protein, 2),
            "fat": round(fat, 2),
            "carbs": round(carbs, 2),
            "fiber": round(fiber, 2),
        },
        "common_portions": [],
    }

wb.close()
print(f"  Parsed {len(foods)} food items.")


# ─────────────────────────────────────────────────────────
# 2. CLASSIFY FOOD GROUPS (name-based approach)
# ─────────────────────────────────────────────────────────
print("\nStep 2: Classifying food groups from food names...")

# Order matters — more specific patterns first
FOOD_GROUP_RULES = [
    # Beverages (check before other categories)
    (["tea,", "tea ", "coffee", "milo", "nescafe", "ovaltine", "horlicks"], "beverages"),
    (["juice", "cordial", "nectar", "squash", "drink"], "beverages"),
    (["coconut water"], "beverages"),
    
    # Cereals & Starches
    (["rice,", "rice ", "fried rice", "biryani", "nasi"], "cereals"),
    (["hopper", "string hopper", "appa", "idiappa"], "cereals"),
    (["pittu"], "cereals"),
    (["bread", "toast", "bun ", "bun,", "roll,", "roll "], "cereals"),
    (["roti", "chapati", "naan", "paratha", "parotta", "godamba"], "cereals"),
    (["noodle", "pasta", "macaroni", "spaghetti", "kottu"], "cereals"),
    (["cereal", "oat", "muesli", "corn flake", "wheat"], "cereals"),
    (["flour", "semolina", "sooji"], "cereals"),
    (["kenda", "porridge", "congee", "kanji"], "cereals"),
    (["kiribath"], "cereals"),
    
    # Protein - Meat
    (["chicken", "poultry"], "protein"),
    (["beef", "mutton", "lamb", "goat", "pork", "venison"], "protein"),
    (["sausage", "bacon", "ham ", "ham,", "salami", "hot dog"], "protein"),
    (["liver", "kidney", "tripe", "offal", "brain", "tongue"], "protein"),
    (["meat"], "protein"),
    
    # Protein - Fish & Seafood
    (["fish", "tuna", "sardine", "mackerel", "anchovy", "sprat"], "protein"),
    (["prawn", "shrimp", "crab", "lobster", "squid", "cuttlefish"], "protein"),
    (["dried fish", "maldive fish"], "protein"),
    (["shark", "ray,", "ray "], "protein"),
    
    # Protein - Eggs
    (["egg,", "egg ", "eggs,", "eggs ", "omelette", "omelet"], "protein"),
    
    # Legumes & Pulses
    (["dhal", "dal ", "dal,", "lentil", "parippu"], "legumes"),
    (["chickpea", "chick pea", "chana"], "legumes"),
    (["mung", "moong", "green gram"], "legumes"),
    (["soy ", "soy,", "soya", "tofu", "tempeh"], "legumes"),
    (["bean", "cowpea", "black gram", "urad"], "legumes"),
    
    # Vegetables
    (["carrot", "potato", "pumpkin", "gourd", "brinjal", "eggplant", "aubergine"], "vegetables"),
    (["tomato", "onion", "leek", "garlic", "ginger,"], "vegetables"),
    (["cabbage", "cauliflower", "broccoli"], "vegetables"),
    (["spinach", "kangkung", "kankun", "gotukola", "centella"], "vegetables"),
    (["murunga", "drumstick", "moringa"], "vegetables"),
    (["kathurumurunga"], "vegetables"),
    (["jack,", "jackfruit", "jak ", "jak,", "breadfruit", "del "], "vegetables"),
    (["sweet potato", "yam", "manioc", "cassava", "innala"], "vegetables"),
    (["capsicum", "pepper,", "chilli", "chili"], "vegetables"),
    (["beetroot", "beet,", "beet "], "vegetables"),
    (["cucumber", "radish", "turnip"], "vegetables"),
    (["mallum", "mallung", "salad", "pickle", "achar"], "vegetables"),
    (["sambol", "pol sambol"], "condiments"),
    (["curry,", "curry ", "tempered", "boiled,", "stir-fr"], "vegetables"),
    (["leaves,", "leaf,", "greens"], "vegetables"),
    (["okra", "ladies finger", "bandakka"], "vegetables"),
    (["ash plantain", "alu kesel"], "vegetables"),
    (["kohila", "lotus"], "vegetables"),
    (["snake gourd", "bitter gourd", "ridge gourd", "bottle gourd", "luffa"], "vegetables"),
    (["wing bean", "winged bean"], "vegetables"),
    
    # Fruits
    (["banana", "plantain"], "fruits"),
    (["mango", "papaya", "pawpaw", "pineapple", "guava"], "fruits"),
    (["apple,", "apple ", "orange,", "orange ", "grape,", "grape "], "fruits"),
    (["watermelon", "melon", "passionfruit", "passion fruit"], "fruits"),
    (["avocado", "rambutan", "mangosteen", "durian", "woodapple"], "fruits"),
    (["jambu", "lovi", "nelli", "amla", "veralu"], "fruits"),
    (["dried fruit", "raisin", "date,", "date ", "prune"], "fruits"),
    (["lime,", "lemon,"], "fruits"),
    
    # Dairy
    (["milk,", "milk ", "curd", "yogurt", "yoghurt"], "dairy"),
    (["cheese", "paneer", "butter,", "butter ", "ghee"], "dairy"),
    (["cream,", "cream ", "ice cream", "icecream"], "dairy"),
    
    # Fats & Oils
    (["oil,", "oil ", "coconut oil", "palm oil", "margarine"], "fats_oils"),
    (["coconut,", "coconut ", "copra", "scraped coconut", "coconut milk"], "fats_oils"),
    (["cashew", "peanut", "groundnut", "almond", "walnut"], "fats_oils"),
    
    # Sweets & Desserts
    (["sugar,", "sugar ", "jaggery", "treacle", "kithul", "honey"], "sweets"),
    (["cake,", "cake ", "pastry", "pudding", "watalappam"], "sweets"),
    (["candy", "toffee", "chocolate", "sweet,", "sweet "], "sweets"),
    (["kavum", "kokis", "aluwa", "aggala", "laddu", "athirasa"], "sweets"),
    (["biscuit", "cookie", "wafer"], "sweets"),
    (["jam,", "jam "], "sweets"),
    (["ice cream", "icecream"], "sweets"),
    
    # Condiments & Spices
    (["sambol", "chutney", "sauce", "ketchup", "mayonnaise"], "condiments"),
    (["salt,", "salt ", "pepper,", "spice", "curry powder", "turmeric"], "condiments"),
    (["cinnamon", "clove", "cardamom", "mustard seed", "cumin", "coriander seed"], "condiments"),
    (["vinegar", "soy sauce"], "condiments"),
    
    # Snack / Short-eats
    (["patties", "pattie", "cutlet", "wade", "vadai", "isso"], "snacks"),
    (["murukku", "chips,", "chips ", "crisp"], "snacks"),
    (["short eat", "samosa", "spring roll", "roll,"], "snacks"),
]

group_counts = {}
for code, food in foods.items():
    name_lower = food["name"].lower()
    matched = False
    for keywords, group in FOOD_GROUP_RULES:
        for kw in keywords:
            if kw in name_lower:
                food["food_group"] = group
                group_counts[group] = group_counts.get(group, 0) + 1
                matched = True
                break
        if matched:
            break
    if not matched:
        food["food_group"] = "other"
        group_counts["other"] = group_counts.get("other", 0) + 1

print(f"  Food group distribution:")
for g, c in sorted(group_counts.items(), key=lambda x: -x[1]):
    print(f"    {g}: {c}")


# ─────────────────────────────────────────────────────────
# 3. TAG: dietary preference + alcohol filter
# ─────────────────────────────────────────────────────────
print("\nStep 3: Tagging dietary preferences & filtering alcohol...")

# Precise alcohol detection — match only actual alcoholic beverages
ALCOHOL_EXACT_NAMES = {
    "beer", "liquor", "wine", "spirits", "cocktail mix",
}
ALCOHOL_STARTS_WITH = [
    "liquor,", "beer,", "wine,", "cocktail",
]

# Non-vegetarian food groups
NON_VEG_FOOD_GROUPS = {"protein"}  # Our protein group is all meat/fish/eggs

# Additional non-veg keywords for items not yet in protein group
NON_VEG_KEYWORDS = [
    "chicken", "fish", "beef", "pork", "mutton", "lamb", "goat",
    "prawn", "shrimp", "crab", "lobster", "squid", "cuttlefish",
    "tuna", "sardine", "mackerel", "anchovy", "sprat",
    "sausage", "bacon", "ham", "meat", "liver", "kidney",
    "egg,", "egg ", "eggs,", "eggs ", "omelette",
    "dried fish", "maldive fish", "shark", "ray,", "ray ",
    "duck", "turkey", "venison", "rabbit",
]

stats = {"veg": 0, "non_veg": 0, "alcohol": 0}

for code, food in foods.items():
    name_lower = food["name"].lower().strip()
    
    # Check alcohol — precise matching
    is_alcohol = False
    if name_lower in ALCOHOL_EXACT_NAMES:
        is_alcohol = True
    elif any(name_lower.startswith(p) for p in ALCOHOL_STARTS_WITH):
        is_alcohol = True
    elif "liquor" in name_lower or ("beer" == name_lower):
        is_alcohol = True
    # Catch remaining: items with ABV in name
    elif "abv" in name_lower or "% alcohol" in name_lower:
        is_alcohol = True
    
    if is_alcohol:
        food["recommendable"] = False
        food["dietary_tags"] = ["non_vegetarian"]
        stats["alcohol"] += 1
        continue
    
    # Check non-vegetarian
    is_non_veg = food["food_group"] in NON_VEG_FOOD_GROUPS
    if not is_non_veg:
        for kw in NON_VEG_KEYWORDS:
            if kw in name_lower:
                is_non_veg = True
                break
    
    if is_non_veg:
        food["dietary_tags"] = ["non_vegetarian"]
        stats["non_veg"] += 1
    else:
        food["dietary_tags"] = ["vegetarian"]
        stats["veg"] += 1

print(f"  Vegetarian: {stats['veg']}")
print(f"  Non-Vegetarian: {stats['non_veg']}")
print(f"  Alcohol (non-recommendable): {stats['alcohol']}")


# ─────────────────────────────────────────────────────────
# 4. TAG MEAL TYPES
# ─────────────────────────────────────────────────────────
print("\nStep 4: Tagging meal types...")

BREAKFAST_KEYWORDS = [
    "hopper", "string hopper", "pittu", "roti", "bread", "toast",
    "milk,", "milk ", "yogurt", "yoghurt", "curd",
    "egg,", "egg ", "omelette", "porridge", "kiribath", "kenda",
    "pancake", "bun ", "bun,", "jam,", "jam ", "butter,", "butter ",
    "tea,", "tea ", "coffee", "milo", "nescafe", "horlicks", "ovaltine",
    "cereal", "oat", "muesli", "corn flake",
    "cheese", "godamba",
]
LUNCH_DINNER_KEYWORDS = [
    "rice,", "rice ", "fried rice", "biryani",
    "curry,", "curry ", "dhal", "dal ", "dal,", "lentil", "parippu",
    "chicken", "fish", "meat", "beef", "pork", "prawn", "crab",
    "noodle", "kottu", "devilled",
    "sambol", "mallum", "mallung", "tempered",
    "potato", "pumpkin", "brinjal", "gourd", "jackfruit",
    "carrot", "cabbage", "bean", "onion", "tomato",
    "boiled,", "fried,", "stir-fr",
]
SNACK_KEYWORDS = [
    "biscuit", "cookie", "cake,", "cake ", "sweet", "candy", "chocolate",
    "chips,", "chips ", "crisp", "murukku", "kavum", "kokis", "aluwa",
    "fruit", "banana", "mango", "papaya", "pineapple", "apple,", "apple ",
    "orange,", "orange ", "juice", "nut,", "nut ", "cashew", "peanut",
    "ice cream", "pudding", "patties", "cutlet", "roll,",
    "samosa", "spring roll", "wade", "vadai",
    "short eat", "pastry", "wafer",
]

for code, food in foods.items():
    if not food["recommendable"]:
        continue
    
    name_lower = food["name"].lower()
    group = food["food_group"]
    meal_types = set()
    
    if any(kw in name_lower for kw in BREAKFAST_KEYWORDS):
        meal_types.add("breakfast")
    if any(kw in name_lower for kw in LUNCH_DINNER_KEYWORDS):
        meal_types.add("lunch")
        meal_types.add("dinner")
    if any(kw in name_lower for kw in SNACK_KEYWORDS):
        meal_types.add("snack")
    
    # Group-based defaults if no keyword matched
    if not meal_types:
        DEFAULT_MEAL_TYPES = {
            "cereals": {"breakfast", "lunch", "dinner"},
            "protein": {"lunch", "dinner"},
            "legumes": {"lunch", "dinner"},
            "vegetables": {"lunch", "dinner"},
            "fruits": {"breakfast", "snack"},
            "dairy": {"breakfast", "snack"},
            "fats_oils": {"breakfast", "lunch", "dinner"},
            "sweets": {"snack"},
            "snacks": {"snack"},
            "condiments": {"breakfast", "lunch", "dinner"},
            "beverages": {"breakfast", "snack"},
            "other": {"breakfast", "lunch", "dinner", "snack"},
        }
        meal_types = DEFAULT_MEAL_TYPES.get(group, {"breakfast", "lunch", "dinner", "snack"})
    
    food["meal_types"] = sorted(list(meal_types))

mt_counts = {"breakfast": 0, "lunch": 0, "dinner": 0, "snack": 0}
for food in foods.values():
    for mt in food.get("meal_types", []):
        if mt in mt_counts:
            mt_counts[mt] += 1
print(f"  Meal type coverage: {mt_counts}")


# ─────────────────────────────────────────────────────────
# 5. PARSE PORTION CONVERSION TABLE
# ─────────────────────────────────────────────────────────
print("\nStep 5: Parsing Portion Conversion Table...")
pc_path = os.path.join(BASE, "Sri Lanka Portion Conversion Estimation Table_20240418.xlsx")
wb = openpyxl.load_workbook(pc_path, data_only=True)
ws = wb["Portion size conversions"]

portions_all = {}
for row in range(2, ws.max_row + 1):
    code = ws.cell(row, 1).value
    unit = ws.cell(row, 3).value
    conv_factor = ws.cell(row, 4).value
    if not code or not unit: continue
    code = str(code).strip()
    unit = str(unit).strip()
    grams = safe_float(conv_factor) if conv_factor else 1.0
    if code not in portions_all:
        portions_all[code] = []
    portions_all[code].append({"unit": unit, "grams": round(grams, 2)})
wb.close()

# Merge into foods
portions_matched = 0
for code, food in foods.items():
    if code in portions_all:
        portions = portions_all[code]
        cal_per_g = food["per_100g"]["calories"] / 100.0
        for p in portions:
            p["calories"] = round(p["grams"] * cal_per_g, 1)
        food["common_portions"] = portions[:5]
        portions_matched += 1
print(f"  Matched portions to {portions_matched}/{len(foods)} foods.")


# ─────────────────────────────────────────────────────────
# 6. PARSE RECIPES DATABASE
# ─────────────────────────────────────────────────────────
print("\nStep 6: Parsing Recipes Database...")
rec_path = os.path.join(BASE, "Sri Lanka Standard Recipes Database_20240514.xlsx")
wb = openpyxl.load_workbook(rec_path, data_only=True)
ws = wb["Recipes table"]

recipes = {}
for row in range(2, ws.max_row + 1):
    recipe_code = ws.cell(row, 1).value
    recipe_name = ws.cell(row, 2).value
    ing_code    = ws.cell(row, 3).value
    ing_name    = ws.cell(row, 5).value
    ing_frac    = ws.cell(row, 6).value
    if not recipe_code or not recipe_name: continue
    recipe_code = str(recipe_code).strip()
    recipe_name = str(recipe_name).strip()
    if recipe_code not in recipes:
        recipes[recipe_code] = {"id": recipe_code, "name": recipe_name, "ingredients": []}
    recipes[recipe_code]["ingredients"].append({
        "code": str(ing_code).strip() if ing_code else "",
        "name": str(ing_name).strip() if ing_name else "",
        "fraction_pct": round(safe_float(ing_frac), 2),
    })
wb.close()
print(f"  Parsed {len(recipes)} recipes.")


# ─────────────────────────────────────────────────────────
# 7. WRITE OUTPUT JSON FILES
# ─────────────────────────────────────────────────────────
print("\nStep 7: Writing output files...")

foods_list = list(foods.values())
# Remove internal 'source' field before outputting
for f in foods_list:
    f.pop("source", None)

with open(os.path.join(OUT_DIR, "foods_sri_lanka.json"), "w", encoding="utf-8") as f:
    json.dump(foods_list, f, indent=2, ensure_ascii=False)

with open(os.path.join(OUT_DIR, "recipes_sri_lanka.json"), "w", encoding="utf-8") as f:
    json.dump(list(recipes.values()), f, indent=2, ensure_ascii=False)

with open(os.path.join(OUT_DIR, "portions.json"), "w", encoding="utf-8") as f:
    json.dump(portions_all, f, indent=2, ensure_ascii=False)


# ─────────────────────────────────────────────────────────
# 8. SUMMARY & VERIFICATION
# ─────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("SUMMARY")
print("=" * 60)

# Write detailed verify results to file
out_lines = []
rec_count = sum(1 for f in foods_list if f["recommendable"])
non_rec = sum(1 for f in foods_list if not f["recommendable"])
veg = sum(1 for f in foods_list if "vegetarian" in f["dietary_tags"])
nveg = sum(1 for f in foods_list if "non_vegetarian" in f["dietary_tags"])

out_lines.append(f"Total foods: {len(foods_list)}")
out_lines.append(f"Recommendable: {rec_count}")
out_lines.append(f"Non-recommendable (alcohol): {non_rec}")
out_lines.append(f"Vegetarian: {veg}")
out_lines.append(f"Non-vegetarian: {nveg}")
out_lines.append(f"Recipes: {len(recipes)}")
out_lines.append(f"Foods with portions: {portions_matched}")

out_lines.append(f"\nFood group distribution:")
grp = {}
for f in foods_list:
    g = f["food_group"]
    grp[g] = grp.get(g, 0) + 1
for g, c in sorted(grp.items(), key=lambda x: -x[1]):
    out_lines.append(f"  {g}: {c}")

out_lines.append(f"\nMeal type coverage:")
for mt, c in mt_counts.items():
    out_lines.append(f"  {mt}: {c}")

out_lines.append(f"\nAlcohol items (non-recommendable):")
for f in foods_list:
    if not f["recommendable"]:
        out_lines.append(f"  {f['name']}")

out_lines.append(f"\nSample vegetarian foods:")
veg_samples = [f for f in foods_list if "vegetarian" in f["dietary_tags"] and f["recommendable"]]
for f in veg_samples[:10]:
    out_lines.append(f"  {f['name']}: {f['per_100g']['calories']} kcal | {f['food_group']} | {f['meal_types']}")

out_lines.append(f"\nSample non-vegetarian foods:")
nveg_samples = [f for f in foods_list if "non_vegetarian" in f["dietary_tags"] and f["recommendable"]]
for f in nveg_samples[:10]:
    out_lines.append(f"  {f['name']}: {f['per_100g']['calories']} kcal | {f['food_group']} | {f['meal_types']}")

verify_path = os.path.join(os.path.dirname(OUT_DIR), "..", "verify_results.txt")
with open(r"C:\Users\joshw\OneDrive\Desktop\FYP_Work\verify_results.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(out_lines))

for line in out_lines[:20]:
    print(f"  {line}")
print(f"\n  Full results → verify_results.txt")
print("DONE!")
